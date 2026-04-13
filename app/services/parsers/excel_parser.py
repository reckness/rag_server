"""
Excel 文档解析器
支持 .xlsx 和 .xls 格式，每个 Sheet 作为一个顶级节点，表格内容按行提取
"""
import os
from typing import Dict, Any, List
from .base_parser import BaseDocumentParser


class ExcelParser(BaseDocumentParser):
    """Excel 文档解析器（.xlsx / .xls）"""

    # 单个 Sheet 最大读取行数（防止超大文件 OOM）
    MAX_ROWS = 2000

    def parse(self, file_path: str, doc_name: str = None) -> Dict[str, Any]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        if doc_name is None:
            doc_name = os.path.splitext(os.path.basename(file_path))[0]
        ext = os.path.splitext(file_path)[1].lower()

        # .xls 需要用 xlrd 读取后转换，或提示安装
        if ext == ".xls":
            return self._parse_xls(file_path, doc_name)

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        structure = []
        node_counter = [0]

        def next_id():
            node_counter[0] += 1
            return str(node_counter[0]).zfill(4)

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            rows_data: List[List[str]] = []
            header: List[str] = []
            row_count = 0

            for row in ws.iter_rows(values_only=True):
                if row_count > self.MAX_ROWS:
                    rows_data.append([f"... (超过 {self.MAX_ROWS} 行，已截断)"])
                    break
                cells = [str(c) if c is not None else "" for c in row]
                # 跳过全空行
                if not any(c.strip() for c in cells):
                    row_count += 1
                    continue
                if row_count == 0 and not header:
                    header = cells
                    row_count += 1
                    continue
                rows_data.append(cells)
                row_count += 1

            # 构造表格文本
            text_lines: List[str] = []
            if header:
                text_lines.append(" | ".join(header))
                text_lines.append("-" * max(len(" | ".join(header)), 20))
            for row in rows_data:
                text_lines.append(" | ".join(row))

            sheet_text = "\n".join(text_lines)

            # 每个 Sheet 对应一个顶级节点
            node = self._make_node(
                title=sheet_name,
                text=sheet_text,
                node_id=next_id(),
                start_index=sheet_idx + 1,
                end_index=sheet_idx + 1,
            )
            structure.append(node)

        wb.close()
        return {"doc_name": doc_name, "structure": structure}

    # ------------------------------------------------------------------
    def _parse_xls(self, file_path: str, doc_name: str) -> Dict[str, Any]:
        """使用 xlrd 解析老版 .xls 文件"""
        try:
            import xlrd
        except ImportError:
            raise ImportError("解析 .xls 文件请安装 xlrd: pip install xlrd")

        wb = xlrd.open_workbook(file_path)
        structure = []
        node_counter = [0]

        def next_id():
            node_counter[0] += 1
            return str(node_counter[0]).zfill(4)

        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            text_lines: List[str] = []
            header_written = False

            for row_idx in range(min(ws.nrows, self.MAX_ROWS + 1)):
                cells = [str(ws.cell_value(row_idx, col)) for col in range(ws.ncols)]
                if not any(c.strip() for c in cells):
                    continue
                if not header_written:
                    text_lines.append(" | ".join(cells))
                    text_lines.append("-" * 40)
                    header_written = True
                else:
                    text_lines.append(" | ".join(cells))

            if ws.nrows > self.MAX_ROWS + 1:
                text_lines.append(f"... (超过 {self.MAX_ROWS} 行，已截断)")

            node = self._make_node(
                title=ws.name,
                text="\n".join(text_lines),
                node_id=next_id(),
                start_index=sheet_idx + 1,
                end_index=sheet_idx + 1,
            )
            structure.append(node)

        return {"doc_name": doc_name, "structure": structure}
